"""market_fetch.py - Fetch monthwise market-share data from Google Sheets.

Reads the "Market Share" tab (gid 1519626691) of the ARMCL market tracker
spreadsheet via the Drive export endpoint (the file is an Office xlsx, so it
cannot be read through the Sheets API). Cached in memory to avoid hammering
Google on every request.

Usage (as module):
    from market_fetch import fetch_market_trend

Returns {"items": [...], "updated_at": ...} or {"error": ...}.

Config:
  MARKET_SHEET_FILE  env / default "1NDWiW6q1PuykQ2uuNLcMuyU_90tVSBqLARlQxiaPgss"
  MARKET_SHEET_GID    env / default "1519626691"
  MARKET_SHEET_TTL    env (seconds, default 3600)
  GOOGLE_TOKEN_JSON   env (optional) raw JSON of the Google credentials/token
                      file. When set, the token is read from the environment
                      instead of database/token.json (serverless / Vercel).
"""

from __future__ import annotations

import io
import json
import os
import time
import urllib.request
from datetime import datetime
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "database"
TOKEN_FILE = DATA_DIR / "token.json"

SHEET_FILE = os.getenv("MARKET_SHEET_FILE", "1NDWiW6q1PuykQ2uuNLcMuyU_90tVSBqLARlQxiaPgss")
SHEET_GID = os.getenv("MARKET_SHEET_GID", "1519626691")
TTL = int(os.getenv("MARKET_SHEET_TTL", "3600"))

SALES_SHEET_FILE = os.getenv("SALES_SHEET_FILE", "1vPlcijsZkj4p6ZHmzg7jEAutNrW5l2YKlbNUtgXkNbI")
SALES_SHEET_GID = os.getenv("SALES_SHEET_GID", "990537426")

_cache: dict = {"data": None, "at": 0, "token": None, "sales": None, "sales_at": 0}


def _token_data() -> dict | None:
    raw = os.getenv("GOOGLE_TOKEN_JSON", "").strip()
    if raw:
        try:
            return json.loads(raw)
        except Exception:
            return None
    if not TOKEN_FILE.exists():
        return None
    try:
        return json.loads(TOKEN_FILE.read_text(encoding="utf-8"))
    except Exception:
        return None


def _refresh(data: dict) -> dict:
    try:
        from google.oauth2.credentials import Credentials
        from google.auth.transport.requests import Request

        creds = Credentials.from_authorized_user_info(data)
        creds.refresh(Request())
        return json.loads(creds.to_json())
    except Exception:
        return data


def _load_token(force_refresh: bool = False) -> str | None:
    data = _token_data()
    if not data:
        return None
    token = data.get("token")
    expiry = data.get("expiry")
    if not token:
        return None
    needs = force_refresh
    if not needs and expiry:
        try:
            exp = datetime.fromisoformat(str(expiry).replace("Z", "+00:00"))
            if datetime.now(exp.tzinfo) >= exp - __import__("datetime").timedelta(minutes=2):
                needs = True
        except Exception:
            needs = True
    if needs and data.get("refresh_token"):
        data = _refresh(data)
        _cache["token"] = data.get("token") or token
        return data.get("token") or _cache["token"]
    _cache["token"] = token
    return token
    try:
        return json.loads(TOKEN_FILE.read_text(encoding="utf-8")).get("token")
    except Exception:
        return None


def _export_xlsx(sheet_file: str, gid: str) -> bytes:
    token = _load_token()
    if not token:
        raise RuntimeError("no Google token configured")
    url = f"https://docs.google.com/spreadsheets/d/{sheet_file}/export?format=xlsx&gid={gid}"
    req = urllib.request.Request(url, headers={"Authorization": "Bearer " + token})
    with urllib.request.urlopen(req, timeout=90) as r:
        return r.read()


def _parse(raw: bytes) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), data_only=True)
    headers = None
    items: list[dict] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            if not any(v is not None for v in row):
                continue
            vals = ["" if v is None else v for v in row]
            if headers is None:
                # first nonempty row is the header
                headers = vals
                continue
            if str(vals[0]).strip().lower() == "avg":
                continue
            month = vals[0]
            # normalize month to a label
            if isinstance(month, datetime):
                # sheet encodes year as day: e.g. 2026-01-25 => Jan 2025, 2026-01-26 => Jan 2026
                year = 2000 + (month.day % 100) if month.day >= 1 else month.year
                label = f"{year:04d}-{month.month:02d}"
            else:
                label = str(month).strip()
            try:
                def num(x):
                    try:
                        return round(float(x), 2)
                    except Exception:
                        return None
                g = vals[6] if len(vals) > 6 else None
                if g is not None:
                    try:
                        g = round(float(g) * 100, 2)
                    except Exception:
                        g = None
                items.append({
                    "month": label,
                    "shah": num(vals[1]),
                    "crown": num(vals[2]),
                    "nde": num(vals[3]),
                    "basundhara": num(vals[4]),
                    "akij": num(vals[5]),
                    "market_share_akij": g,
                })
            except IndexError:
                continue
    return items


def get_market_trend(force: bool = False) -> dict:
    now = time.time()
    if not force and _cache["data"] is not None and (now - _cache["at"]) < TTL:
        return _cache["data"]
    try:
        raw = _export_xlsx(SHEET_FILE, SHEET_GID)
        items = _parse(raw)
        data = {
            "items": items,
            "source": "Google Sheet",
            "sheet": SHEET_FILE,
            "gid": SHEET_GID,
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        _cache["data"] = data
        _cache["at"] = now
        return data
    except Exception as e:  # noqa: BLE001
        return {"error": str(e), "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}


def _finance_monthly_target(month_label: str) -> float | None:
    """Monthly sales target (Sales Qty CFT) from the Finance Budget store.

    month_label forms like "Aug'26" are mapped to the canonical YYYY-MM key
    used in database/finance_budget.json (e.g. 2026-08). Returns None when the
    finance data is missing so callers can fall back to the statement target.
    """
    import json as _json

    try:
        store = Path(DATA_DIR / "finance_budget.json")
        if not store.exists():
            return None
        rows = _json.loads(store.read_text(encoding="utf-8"))
    except Exception:
        return None
    if not isinstance(rows, list):
        return None

    # map "Aug'26" -> "2026-08"
    months = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
              "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}
    mkey = None
    low = (month_label or "").lower().replace("'", "").replace(" ", "")
    for name, m in months.items():
        if low.startswith(name):
            nums = [w for w in low if w.isdigit()]
            yy = int("".join(nums[-2:])) if len(nums) >= 2 else 26
            yyyy = 2000 + yy
            mkey = f"{yyyy:04d}-{m:02d}"
            break
    if not mkey:
        # already a YYYY-MM form?
        import re
        mm = re.match(r"(\d{4})-(\d{2})", month_label or "")
        if mm:
            mkey = mm.group(0)

    if not mkey:
        return None
    for r in rows:
        if str(r.get("month", "")) == mkey:
            try:
                qty = float(r.get("sales_qty_cft"))
                return qty
            except (TypeError, ValueError):
                return None
    return None


def get_sales_status(force: bool = False) -> dict:
    """MTD sales achievement from the ARMCL-Aug26 customer-statement sheet.

    Reads the 'Information At a Glance' key/value pairs (column P = label,
    column Q = value) and the month header. Used to reconcile Sales
    Performance vs DWH actuals.

    Monthly sales TARGET is taken from the FINANCE BUDGET sheet (Source A)
    when available for the current month (source-of-truth precedence).
    """
    now = time.time()
    if not force and _cache["sales"] is not None and (now - _cache["sales_at"]) < TTL:
        return _cache["sales"]
    try:
        raw = _export_xlsx(SALES_SHEET_FILE, SALES_SHEET_GID)
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.worksheets[0]

        sheet_label = ws.title
        month_label = ""
        labels: dict[str, float] = {}
        for row in ws.iter_rows(values_only=True):
            if not any(v is not None for v in row):
                continue
            # capture the "Month Name: ..." header if present
            first = row[0]
            if isinstance(first, str) and first.startswith("Month Name:"):
                month_label = first.split(":", 1)[1].strip()
            label = row[15] if len(row) > 15 else None
            value = row[16] if len(row) > 16 else None
            if isinstance(label, str) and label.strip() and value is not None:
                try:
                    labels[label.strip()] = float(value)
                except (TypeError, ValueError):
                    pass

        def _f(label: str) -> float | None:
            v = labels.get(label)
            return v

        monthly_target = _f("Monthly Target")
        mtd_sales = _f("Sales Till Date")
        achievement = _f("Achiv % till date")
        present_ads = _f("Present ADS")
        logical_sales = _f("Logical Sales till date")
        remaining = _f("Remaining Sales")
        days_consumed = _f("Day's consumed")
        days_remaining = _f("Days Remaining")
        rads = _f("RADS")

        # Source-of-truth override: monthly sales target from the FINANCE BUDGET
        # sheet (Source A) when available for the current month.
        finance_target = _finance_monthly_target(month_label)
        if finance_target is not None:
            monthly_target = finance_target
            target_source = "Finance Budget (Source A)"
        else:
            target_source = "Customer Statement (E2)"

        achievement_pct = (achievement * 100.0) if achievement is not None else None
        if monthly_target and mtd_sales and achievement_pct is None:
            achievement_pct = round((mtd_sales / monthly_target) * 100.0, 2)
        elif monthly_target and mtd_sales:
            # recompute against the authoritative target
            achievement_pct = round((mtd_sales / monthly_target) * 100.0, 2)

        data = {
            "source": "Google Sheet",
            "sheet": SALES_SHEET_FILE,
            "gid": SALES_SHEET_GID,
            "tab": sheet_label,
            "month": month_label,
            "monthly_target": monthly_target,
            "monthly_target_source": target_source,
            "mtd_sales": mtd_sales,
            "present_ads": present_ads,
            "logical_sales_till_date": logical_sales,
            "remaining_sales": remaining,
            "achievement_pct": round(achievement_pct, 2) if achievement_pct is not None else None,
            "days_consumed": days_consumed,
            "days_remaining": days_remaining,
            "rads": rads,
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        _cache["sales"] = data
        _cache["sales_at"] = now
        return data
    except Exception as e:  # noqa: BLE001
        # Even if the statement is unreachable, still surface the authoritative
        # finance-budget target for the current month when available.
        fallback = {"error": str(e), "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
        try:
            tgt = _finance_monthly_target(datetime.now().strftime("%b'%y"))
            if tgt is not None:
                fallback["monthly_target"] = tgt
                fallback["monthly_target_source"] = "Finance Budget (Source A)"
                fallback["mtd_sales"] = None
                fallback["achievement_pct"] = None
        except Exception:
            pass
        return fallback


def refresh_sales() -> dict:
    return get_sales_status(force=True)


def refresh() -> dict:
    return {
        "market": get_market_trend(force=True),
        "sales": get_sales_status(force=True),
    }


if __name__ == "__main__":
    print(json.dumps(refresh(), ensure_ascii=False, indent=2))