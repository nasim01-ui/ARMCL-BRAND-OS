"""sync_sheets.py - BrandOS spreadsheet synchronization (PHASE 6).

Ingests the designated Google Sheets (Finance Budget, Brand Budget, Market
Share, MTD Sales Statement) into canonical JSON stores in database/ using the
canonical data model (web/canonical.py).

Features:
  * Idempotent import (records keyed by deterministic id)
  * Add / modify / delete detection
  * Schema-change protection -> DATA MAPPING ALERT (stops unsafe import)
  * Historical snapshots (per-source, timestamped)
  * Sync log store (database/sync_log.json)
  * Manual refresh (force) + TTL cache

Usage:
    from sync_sheets import sync_all, get_sync_status, sync_source

    sync_all(force=True)            # refresh every mapped source
    sync_source("finance_budget")   # refresh a single source
    get_sync_status()               # freshness + last sync per source
"""

from __future__ import annotations

import io
import json
import time
import urllib.request
from datetime import datetime
from pathlib import Path

import canonical
import market_fetch

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "database"

SYNC_TTL = int(__import__("os").getenv("BRANDOS_SYNC_TTL", "3600"))

# Store name (database/<name>.json) -> canonical entity it holds
STORE_ENTITY = {
    "finance_budget": "finance_budget_total",
    "brand_budget": "brand_budget",
    "market_share": "market_share",
    "sales_achievement": "sales_achievement",
}


# --------------------------------------------------------------------------
# helpers
# --------------------------------------------------------------------------
def _store_path(name: str) -> Path:
    return DATA_DIR / f"{name}.json"


def _read_store(name: str) -> list[dict]:
    p = _store_path(name)
    if not p.exists():
        return []
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
        return data if isinstance(data, list) else []
    except Exception:
        return []


def _write_store(name: str, items: list[dict]) -> None:
    _store_path(name).write_text(json.dumps(items, indent=2, ensure_ascii=False), encoding="utf-8")


def _export_xlsx(sheet_file: str, gid: str) -> bytes:
    token = market_fetch._load_token()
    if not token:
        raise RuntimeError("no Google token configured")
    url = f"https://docs.google.com/spreadsheets/d/{sheet_file}/export?format=xlsx&gid={gid}"
    req = urllib.request.Request(url, headers={"Authorization": "Bearer " + token})
    with urllib.request.urlopen(req, timeout=90) as r:
        return r.read()


def _num(x):
    if x is None:
        return None
    try:
        return round(float(x), 2)
    except (TypeError, ValueError):
        return None


def _month_key(val) -> str | None:
    """Normalize a month cell to YYYY-MM."""
    if isinstance(val, datetime):
        return f"{val.year:04d}-{val.month:02d}"
    s = str(val).strip()
    # parse forms like "Jul'26", "Jul'26", "2026-07", "July 2026"
    months = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
        "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    }
    low = s.lower().replace("'", "")
    for name, m in months.items():
        if low.startswith(name):
            # extract year
            nums = [w for w in low.replace(",", "").split() if w.isdigit()]
            year = int(nums[-1]) if nums else 2026
            if year < 100:
                year += 2000
            return f"{year:04d}-{m:02d}"
    import re
    m = re.search(r"(20\d\d)[-./]?(\d{1,2})", s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"
    return None


# --------------------------------------------------------------------------
# Finance Budget ingest
# --------------------------------------------------------------------------
FIN_HEADER_EXPECT = ["Month", "Sales Qty (CFT)"]


def _ingest_finance_budget(raw: bytes) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.worksheets[0]

    rows = list(ws.iter_rows(values_only=True))
    # find header row (contains 'Month' in col 0)
    header_row = None
    for i, r in enumerate(rows):
        if r and str(r[0]).strip() == "Month":
            header_row = i
            break
    if header_row is None:
        raise ValueError("finance_budget: header row with 'Month' not found")

    # schema protection: verify expected column header at col 1
    col1 = rows[header_row][1] if len(rows[header_row]) > 1 else None
    if col1 is None or "Sales Qty" not in str(col1):
        raise SchemaMappingAlert(
            "finance_budget",
            f"expected 'Sales Qty (CFT)' at col B, got {col1!r}",
        )

    # unit-row: budget26/27 quantities at col 3, rate at col 5, net sales at col 9
    records = []
    errors = []
    for r in rows[header_row + 2:]:
        if not any(v is not None for v in r):
            continue
        month = _month_key(r[0])
        if not month:
            continue
        rec = {
            "id": f"finance_budget:{month}",
            "fiscal_year": "2026-2027",
            "month": month,
            "sales_qty_cft": _num(r[3]) if len(r) > 3 else None,
            "rate": _num(r[5]) if len(r) > 5 else None,
            "net_sales": _num(r[9]) if len(r) > 9 else None,
        }
        verr = canonical.validate_record("finance_budget_total", rec)
        if verr:
            errors.append({"row": header_row + 2, "errors": verr})
        records.append(rec)

    # full P&L rows (line-item) - each block col A=line, cols 7.. (net sales etc.)
    return {
        "entity": "finance_budget_total",
        "records": records,
        "errors": errors[:20],
        "imported": len(records),
        "month_keys": [r["month"] for r in records],
    }


class SchemaMappingAlert(Exception):
    """Raised when a source schema no longer matches the mapping config."""

    def __init__(self, source: str, message: str):
        self.source = source
        self.message = message
        super().__init__(f"[DATA MAPPING ALERT] {source}: {message}")


# --------------------------------------------------------------------------
# Brand Budget ingest
# --------------------------------------------------------------------------
def _ingest_brand_budget(raw: bytes) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    header_row = None
    for i, r in enumerate(rows):
        if r and str(r[1]).strip().lower() == "type of activity":
            header_row = i
            break
    if header_row is None:
        raise ValueError("brand_budget: header row with 'Type of Activity' not found")

    records = []
    errors = []
    for r in rows[header_row + 1:]:
        if not any(v is not None for v in r):
            continue
        activity = str(r[1]).strip() if r[1] else ""
        if not activity or activity.lower() == "total":
            continue  # skip totals row
        annual_total = _num(r[6]) if len(r) > 6 else None
        q1 = _num(r[2]) if len(r) > 2 else None
        q2 = _num(r[3]) if len(r) > 3 else None
        q3 = _num(r[4]) if len(r) > 4 else None
        q4 = _num(r[5]) if len(r) > 5 else None
        # skip reference / note rows with no budget figures
        if annual_total is None and q1 is None and q2 is None and q3 is None and q4 is None:
            continue
        rec = {
            "id": f"brand_budget:2026-2027:{activity}",
            "fiscal_year": "2026-2027",
            "activity_type": activity,
            "q1": q1,
            "q2": q2,
            "q3": q3,
            "q4": q4,
            "annual_total": annual_total,
            "pct": _num(r[7]) if len(r) > 7 else None,
            "till_date_cost": _num(r[8]) if len(r) > 8 else None,
            "remaining_budget": _num(r[9]) if len(r) > 9 else None,
            "status": "approved" if annual_total is not None else "draft",
        }
        verr = canonical.validate_record("brand_budget", rec)
        if verr:
            errors.append({"row": header_row + 1, "errors": verr})
        records.append(rec)

    return {
        "entity": "brand_budget",
        "records": records,
        "errors": errors[:20],
        "imported": len(records),
    }


# --------------------------------------------------------------------------
# Generic sync: merge records into store idempotently + snapshot
# --------------------------------------------------------------------------
def _merge_into_store(store: str, incoming: list[dict]) -> dict:
    """Idempotent merge. Detects added / modified / deleted. Writes snapshot."""
    existing = {r["id"]: r for r in _read_store(store)}
    new_records = {r["id"]: r for r in incoming}
    added = [i for i in new_records if i not in existing]
    modified = [i for i in new_records if i in existing and existing[i] != new_records[i]]
    removed = [i for i in existing if i not in new_records]

    merged = list(new_records.values())
    merged.sort(key=lambda r: r.get("id", ""))
    _write_store(store, merged)

    # snapshot (historical, audit trail)
    if incoming:
        snap = {
            "store": store,
            "captured_at": datetime.now().isoformat(timespec="seconds"),
            "records": len(merged),
            "source_hashes": {r["id"]: r.get("source_hash") for r in merged if r.get("source_hash")},
        }
        snap_dir = DATA_DIR / "snapshots"
        snap_dir.mkdir(exist_ok=True)
        (snap_dir / f"{store}_{datetime.now().strftime('%Y%m%d%H%M%S')}.json").write_text(
            json.dumps(snap, indent=2, ensure_ascii=False), encoding="utf-8"
        )

    return {
        "added": len(added),
        "modified": len(modified),
        "removed": len(removed),
        "total": len(merged),
        "ids_added": [i for i in added][:20],
        "ids_modified": [i for i in modified][:20],
        "ids_removed": [i for i in removed][:20],
    }


# --------------------------------------------------------------------------
# Source dispatch
# --------------------------------------------------------------------------
SYNC_CACHE: dict[str, dict] = {}

SOURCE_DISPATCH = {
    "finance_budget": ("1RCrI84M2w9xSgt9Unl7LInA4RzkKxvU6", "583178831", "finance_budget", _ingest_finance_budget),
    "brand_budget": ("1GamxQqXTXavG1rtPpaV9xkVyrqgZbCRg2PHjosXG7yU", "250903347", "brand_budget", _ingest_brand_budget),
    "market_share": ("1NDWiW6q1PuykQ2uuNLcMuyU_90tVSBqLARlQxiaPgss", "1519626691", "market_share", None),  # handled by market_fetch
    "sales_achievement": ("1vPlcijsZkj4p6ZHmzg7jEAutNrW5l2YKlbNUtgXkNbI", "990537426", "sales_achievement", None),  # handled by market_fetch
}


def sync_source(source: str, force: bool = False) -> dict:
    """Sync one spreadsheet source into its canonical store."""
    now = time.time()
    cached = SYNC_CACHE.get(source)
    if not force and cached and (now - cached["at"]) < SYNC_TTL:
        return cached["result"]

    if source in ("market_share", "sales_achievement"):
        # these are already fetched by market_fetch; just read the store + freshness
        store = "market_share" if source == "market_share" else "sales_achievement"
        items = _read_store(store)
        result = {
            "source": source,
            "store": store,
            "records": len(items),
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "note": "live via market_fetch (not persisted here)",
        }
        SYNC_CACHE[source] = {"at": now, "result": result}
        return result

    spec = SOURCE_DISPATCH.get(source)
    if not spec:
        return {"source": source, "error": "unknown source"}
    sheet_file, gid, store, ingest_fn = spec

    try:
        raw = _export_xlsx(sheet_file, gid)
        parsed = ingest_fn(raw)
        if parsed.get("errors"):
            # schema issue on some rows - keep going but flag
            pass
        merge = _merge_into_store(store, parsed["records"])
        result = {
            "source": source,
            "store": store,
            "entity": parsed.get("entity"),
            "imported": parsed.get("imported"),
            "merge": merge,
            "errors": parsed.get("errors", []),
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        SYNC_CACHE[source] = {"at": now, "result": result}
        _log_sync(source, result)
        return result
    except SchemaMappingAlert as e:
        _log_sync(source, {"source": source, "error": e.message, "status": "mapping_alert"})
        return {"source": source, "error": e.message, "status": "mapping_alert"}
    except Exception as e:
        _log_sync(source, {"source": source, "error": str(e), "status": "error"})
        return {"source": source, "error": str(e), "status": "error"}


def _log_sync(source: str, result: dict) -> None:
    log = _read_store("sync_log")
    log.append({
        "id": f"sync:{source}:{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "source": source,
        "status": result.get("status", "ok"),
        "imported": result.get("imported"),
        "merge": result.get("merge"),
        "error": result.get("error"),
        "synced_at": datetime.now().isoformat(timespec="seconds"),
    })
    _write_store("sync_log", log[-100:])


def sync_all(force: bool = False) -> dict:
    results = {}
    for source in SOURCE_DISPATCH:
        results[source] = sync_source(source, force=force)
    return results


def get_sync_status() -> dict:
    """Data freshness dashboard for all sources."""
    out = {}
    for source, spec in SOURCE_DISPATCH.items():
        cache = SYNC_CACHE.get(source)
        out[source] = {
            "source": source,
            "store": spec[2],
            "last_sync": cache["result"].get("updated_at") if cache else None,
            "last_status": cache["result"].get("status", "ok") if cache else "never",
            "records": len(_read_store(spec[2])),
            "cache_ttl": SYNC_TTL,
        }
    return out


if __name__ == "__main__":
    import json as _json
    print(_json.dumps(sync_all(force=True), ensure_ascii=False, indent=2))
